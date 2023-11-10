#Função para colocar senha no excel 
import pandas as pd
from openpyxl import load_workbook

#Definindo o arquivo com uma senha 
def crypto_excel(path_data, key):
    print('Colocando senha no arquivo...')
    df = pd.read_excel(path_data, engine='openpyxl')
    
    #Escrever os dados no excel sem senha
    with pd.ExcelWriter(path_data, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Carregando o livro de trabalho
    book = load_workbook(path_data)

    # Adicionando proteção à planilha com senha
    for sheet in book.sheetnames:
        ws = book[sheet]
        ws.protection.password = key
        ws.protection.enable()

    # Salvando as alterações no arquivo Excel
    book.save(path_data)

    print('Senha colocada com sucesso!')


